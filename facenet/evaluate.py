# -*- coding: utf-8 -*-
"""
Created on Mon Mar 16 20:50:33 2020

@author: andrea
"""
from tkinter import filedialog as FileDialog
import numpy as np
import os
import matplotlib.pyplot as plt
import scipy.stats as stats
import pickle
import pandas as pd
import openpyxl 



def excel_range(wb,named_range):
    
    dir0 = wb.defined_names[named_range]

    dir = dir0.value.split('!')
    sheet = dir[0].replace("'","")
    rg = dir[1].split(':')
    #r = dir[1].replace('$','').split(':')
    ws = wb[sheet]

    data_rows=[]
    for row in ws[rg[0]:rg[1]]:
        data_cols=[]
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    df = pd.DataFrame(data_rows)
    return(df)

def excel_write(ws,vArray,nrow,ncol):
    
    for i,row in enumerate(vArray):
        for j,val in enumerate(row):               
            ws.cell(row=nrow+i,column=ncol+j).value = val           
       
        
    
   


def cllr(GT,LR):

# replace nan for 1 
# nan means the software didn't find the face
# replace small values for 1e-5 (LLR = -5)
# replace large values for 1e5 (LLR = 5)

   
    LR = np.nan_to_num(LR,nan=1.0)
    LR[LR < 1e-5] = 1e-5
    LR[LR > 1e5] = 1e5
    
    LRp = LR[GT > 0]
    LRd = LR[GT < 0]
    
    Np = np.sum(GT>0)
    Nd = np.sum(GT<0)   
    
    LRp2 = np.log2(1 +1.0/LRp);
    LRd2 = np.log2(1 + LRd);
    
    Cllr = (np.nanmean(LRp2) + np.nanmean(LRd2))/2;
    
    TP = sum(LR[GT>0]> 2)
    TN = sum(LR[GT<0]< .5) 
    FP = sum(LR[GT<0]> 2)
    FN = sum(LR[GT>0]< .5) 
    MCC = (TP*TN - FP*FN)/np.sqrt((TP+FP)*(TP+FN)*(TN+FP)*(TN+FN));
    return([Cllr,TP,TN,FP,FN,MCC])


def LR2PTS( LR ):
# Convert Likelihood Ratio to Proficiency test scores
#Detailed explanation goes here

#%Transforming LR to Proficiency test scores.

#% me aseguro que trabajo con un vector columna.
    LR = LR.flatten()

#%LR2 = LR traduciendo valores menor que 1. me quedo con idx para cambiar el
#%signo después (no ahora)
    LR2 = LR;
    idx = LR2<1
    LR2[idx]= 1.0/LR2[idx]   


#%Genero el vector de scores
    PTS = 0*LR2
    PTS = PTS.astype(int) 

    PTS[LR2>=2]=1
    PTS[LR2>=10]=2
    PTS[LR2>=100]=3
    PTS[LR2>=10000]=4
    PTS[LR2>=1000000]=5

#% y pongo el signo negativo donde procede

    PTS[idx] = -1*PTS[idx]
    return(PTS)

#%% ask for distribution file
file = FileDialog.askopenfilename(
    initialdir=".", 
    filetypes=(
        ("distribution data files","*.dat"),
        ("Ficheros de texto", "*.txt"),
        
    ), 
    title = "Open Distribution file"
)

with open(file,'rb') as f:
    [dif_W,dif_K,same_W,same_K] = pickle.load(f)
    f.close
    
#%% ask for excel file
file_ex = FileDialog.askopenfilename(
    initialdir=".", 
    filetypes=(
        ("excel files","*.xlsx"),
        ("Ficheros de texto", "*.txt"),
        
    ), 
    title = "Open Excel file"
)

# define excel output file name

path = os.path.dirname(file)
filename = os.path.basename(file).split('.')[0]
file_save = path+os.path.sep+filename+'_res.xlsx'

#%% read from excel    

#Open workbook
wb = openpyxl.load_workbook(file_ex,read_only = True)    


sim2017 = excel_range(wb,'facenet_distance_2017')
sim2013 = excel_range(wb,'facenet_distance_2013')
sim2012 = excel_range(wb,'facenet_distance_2012')
sim2011 = excel_range(wb,'facenet_distance_2011')


#%Ground truth
GT_2017 = excel_range(wb,'GT_2017')
GT_2013 = excel_range(wb,'GT_2013')
GT_2012 = excel_range(wb,'GT_2012')
GT_2011 = excel_range(wb,'GT_2011')

GT_2017 = GT_2017.to_numpy().flatten()
GT_2013 = GT_2013.to_numpy().flatten()
GT_2012 = GT_2012.to_numpy().flatten()
GT_2011 = GT_2011.to_numpy().flatten()

wb.close


#% The ranges have 2 columns, we remove the first.
s_2017 = sim2017.drop([0],axis=1).to_numpy().flatten()
s_2013 = sim2013.drop([0],axis=1).to_numpy().flatten()
s_2012 = sim2012.drop([0],axis=1).to_numpy().flatten()
s_2011 = sim2011.drop([0],axis=1).to_numpy().flatten()

#%%
P_a = same_W.pdf(s_2017)
P_d = dif_W.pdf(s_2017)
LR_2017_W = np.true_divide(P_a,P_d)

P_a = same_K.pdf(s_2017)
P_d = dif_K.pdf(s_2017)
LR_2017_K = np.true_divide(P_a,P_d)

Cllr_2017_W = cllr(GT_2017,LR_2017_W)
Cllr_2017_K = cllr(GT_2017,LR_2017_K)

# 2013
P_a = same_W.pdf(s_2013)
P_d = dif_W.pdf(s_2013)
LR_2013_W = np.true_divide(P_a,P_d)

P_a = same_K.pdf(s_2013)
P_d = dif_K.pdf(s_2013)
LR_2013_K = np.true_divide(P_a,P_d)

Cllr_2013_W = cllr(GT_2013,LR_2013_W)
Cllr_2013_K = cllr(GT_2013,LR_2013_K)

#2012
P_a = same_W.pdf(s_2012)
P_d = dif_W.pdf(s_2012)
LR_2012_W = np.true_divide(P_a,P_d)

P_a = same_K.pdf(s_2012)
P_d = dif_K.pdf(s_2012)
LR_2012_K = np.true_divide(P_a,P_d)

Cllr_2012_W = cllr(GT_2012,LR_2012_W)
Cllr_2012_K = cllr(GT_2012,LR_2012_K)

#2011
P_a = same_W.pdf(s_2011)
P_d = dif_W.pdf(s_2011)
LR_2011_W = np.true_divide(P_a,P_d)

P_a = same_K.pdf(s_2011)
P_d = dif_K.pdf(s_2011)
LR_2011_K = np.true_divide(P_a,P_d)

Cllr_2011_W = cllr(GT_2011,LR_2011_W)
Cllr_2011_K = cllr(GT_2011,LR_2011_K)

# LLRs

LLR_2017_W = LR2PTS(LR_2017_W)
LLR_2013_W = LR2PTS(LR_2013_W)
LLR_2012_W = LR2PTS(LR_2012_W)
LLR_2011_W = LR2PTS(LR_2011_W)

LLR_2017_K = LR2PTS(LR_2017_K)
LLR_2013_K = LR2PTS(LR_2013_K)
LLR_2012_K = LR2PTS(LR_2012_K)
LLR_2011_K = LR2PTS(LR_2011_K)

#%% Write in excel
#file_save = 'pepito.xlsx'
#Open workbook
wb = openpyxl.Workbook()



#wb = openpyxl.load_workbook(file_save,read_only = false)

col_header=[
['','Cllr','','','','','',''],
['GT','Distance','LR_Weibull','LR_Kernel','LR_ISO','LLR_Weibull','LLR_Kernel','LLR_ISO'] 
]    #%Row cell array (for column labels)

ws = wb.create_sheet('2017')
excel_write(ws,col_header,1,1)

Cllr_2017 = [[Cllr_2017_W[0],Cllr_2017_K[0],'Cllr_2017_iso']];
excel_write(ws,Cllr_2017,1,3)

A = np.array([GT_2017,s_2017,LR_2017_W,LR_2017_K]).T
excel_write(ws,A,3,1)

A = np.array([LLR_2017_W,LLR_2017_K]).T
excel_write(ws,A,3,6)

#2013
ws = wb.create_sheet('2013')
excel_write(ws,col_header,1,1)

Cllr_2013 = [[Cllr_2013_W[0],Cllr_2013_K[0],'Cllr_2013_iso']];
excel_write(ws,Cllr_2013,1,3)

A = np.array([GT_2013,s_2013,LR_2013_W,LR_2013_K]).T
excel_write(ws,A,3,1)

A = np.array([LLR_2013_W,LLR_2013_K]).T
excel_write(ws,A,3,6)

#2012
ws = wb.create_sheet('2012')
excel_write(ws,col_header,1,1)

Cllr_2012 = [[Cllr_2012_W[0],Cllr_2012_K[0],'Cllr_2012_iso']];
excel_write(ws,Cllr_2012,1,3)

A = np.array([GT_2012,s_2012,LR_2012_W,LR_2012_K]).T
excel_write(ws,A,3,1)

A = np.array([LLR_2012_W,LLR_2012_K]).T
excel_write(ws,A,3,6)

#2011
ws = wb.create_sheet('2011')
excel_write(ws,col_header,1,1)

Cllr_2011 = [[Cllr_2011_W[0],Cllr_2011_K[0],'Cllr_2011_iso']];
excel_write(ws,Cllr_2011,1,3)

A = np.array([GT_2011,s_2011,LR_2011_W,LR_2011_K]).T
excel_write(ws,A,3,1)

A = np.array([LLR_2011_W,LLR_2011_K]).T
excel_write(ws,A,3,6)


wb.save(file_save)
